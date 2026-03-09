#!/usr/bin/env python3
"""
eoat_sweep.py
EOAT Parametric Sweep — FANUC ARC Mate 120iD/12L + Hypertherm XPR300 Torch

Hockey-stick 2-member cantilever frame, PyNiteFEA parametric analysis.
Usage:  python eoat_sweep.py
Requires: pip install PyNiteFEA openpyxl
"""

import math
import time
import sys

try:
    from Pynite import FEModel3D
except ImportError:
    print("ERROR: PyNiteFEA not installed.  Run: pip install PyNiteFEA")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed — Excel output skipped.")

# ══════════════════════════════════════════════════════════════════════════════
# USER CONFIGURATION — CHANGE THESE VALUES
# ══════════════════════════════════════════════════════════════════════════════

# Bracket geometry (mm, degrees)
L1    = 150    # Horizontal leg length: flange to elbow (mm)
L2    = 150    # Angled leg length: elbow to torch (mm)
THETA = 55     # Angle of angled leg below horizontal (degrees)

# Component masses (kg)
TORCH_MASS    = 2.2    # XPR300 torch + receptacle + sleeve
SENSOR_MASS   = 0.5    # IL-300 + garolite + manifold + window + shroud
CABLE_MASS    = 0.86   # Effective cable mass at wrist
FASTENER_MASS = 0.2    # Rivnuts, bolts, shroud sheet, misc

# Dynamic load multiplier
G_FACTOR = 3.0         # 1.0 = static only, 3.0 = typical robot accel

# Deflection limit (mm)
MAX_DEFLECTION = 0.5   # Allowable TCP error budget (mm)

# Gravity sweep increment (degrees)
GRAVITY_STEP = 30      # 12 in-plane orientations at this increment

# ══════════════════════════════════════════════════════════════════════════════
# END USER CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════

G_ACCEL = 9.81  # m/s²

# ── Material Database: E, G in MPa; rho in kg/m³; Fy in MPa ──────────────────
MATERIALS = {
    '6061-T6 Al':  {'E': 68_900,  'G': 26_000, 'rho': 2_700, 'Fy': 276},
    '304 SS':      {'E': 193_000, 'G': 77_000,  'rho': 8_000, 'Fy': 215},
    '4130 Steel':  {'E': 205_000, 'G': 80_000,  'rho': 7_850, 'Fy': 460},
}

# ── Section Sweep Ranges (mm) ─────────────────────────────────────────────────
D_VALUES = [50.8, 76.2, 101.6]       # outer depth:  2", 3", 4"
W_VALUES = [50.8, 76.2, 101.6]       # outer width:  W ≤ D enforced
T_VALUES = [1.59, 2.38, 3.18, 4.76]  # wall thickness: 1/16" … 3/16"

# ── Robot Wrist Limits (FANUC ARC Mate 120iD/12L) ────────────────────────────
ROBOT_LIMITS = {
    'payload_kg':        12.0,
    'j45_moment_Nm':     22.0,
    'j6_moment_Nm':       9.8,
    'j45_inertia_kgm2':   0.65,
    'j6_inertia_kgm2':    0.17,
}


# ─────────────────────────────────────────────────────────────────────────────
# Section properties
# ─────────────────────────────────────────────────────────────────────────────

def rhs_properties(D, W, t):
    """RHS section properties. All dims in mm; returns (A mm², Iy mm⁴, Iz mm⁴, J mm⁴)."""
    A  = 2 * t * (D + W) - 4 * t**2
    Iy = (D**3 * W  - (D - 2*t)**3 * (W - 2*t)) / 12   # strong axis (XY-plane bending)
    Iz = (W**3 * D  - (W - 2*t)**3 * (D - 2*t)) / 12   # weak axis
    J  = 2 * t * (D - t)**2 * (W - t)**2 / (D + W - 2*t)
    return A, Iy, Iz, J


def member_mass_kg(rho_kgm3, A_mm2, L_mm):
    """Member mass [kg]. rho in kg/m³, A in mm², L in mm."""
    return rho_kgm3 * A_mm2 * 1e-6 * L_mm * 1e-3


# ─────────────────────────────────────────────────────────────────────────────
# Load case generation
# ─────────────────────────────────────────────────────────────────────────────

def gravity_cases():
    """
    26 load cases: 12 in-plane × 2 load levels + 1 out-of-plane × 2 load levels.
    Each entry: dict with name, gx, gy, gz [m/s²], lf, alpha, dynamic.
    """
    cases = []
    for alpha in range(0, 360, GRAVITY_STEP):
        ar = math.radians(alpha)
        gx_u = -math.sin(ar)
        gy_u = -math.cos(ar)
        for lf, tag in [(1.0, '1g'), (G_FACTOR, f'{G_FACTOR:.0f}g')]:
            cases.append({
                'name':    f'A{alpha:03d}_{tag}',
                'gx':      gx_u * G_ACCEL * lf,
                'gy':      gy_u * G_ACCEL * lf,
                'gz':      0.0,
                'lf':      lf,
                'alpha':   alpha,
                'dynamic': lf > 1.0,
            })
    # Out-of-plane: gravity in -Z
    for lf, tag in [(1.0, '1g'), (G_FACTOR, f'{G_FACTOR:.0f}g')]:
        cases.append({
            'name':    f'Zdown_{tag}',
            'gx':      0.0,
            'gy':      0.0,
            'gz':      -G_ACCEL * lf,
            'lf':      lf,
            'alpha':   None,   # None = out-of-plane
            'dynamic': lf > 1.0,
        })
    return cases


# ─────────────────────────────────────────────────────────────────────────────
# FEA analysis
# ─────────────────────────────────────────────────────────────────────────────

def run_analysis(mat_name, mat, D, W, t, cases):
    """
    Build one Pynite FEModel3D for this material+section, run all 26 gravity
    load cases in a single solve, extract worst-case results.

    Returns (result_dict, None) on success, (None, error_str) on failure.
    """
    A, Iy, Iz, J = rhs_properties(D, W, t)
    E   = mat['E']
    G   = mat['G']
    rho = mat['rho']
    Fy  = mat['Fy']

    # Frame geometry
    theta_rad = math.radians(THETA)
    x_mid = L1 / 2.0
    x_n2  = float(L1)
    x_n3  = L1 + L2 * math.cos(theta_rad)
    y_n3  = -L2 * math.sin(theta_rad)

    # Member lengths
    L_m1a = x_mid
    L_m1b = x_n2 - x_mid
    L_m2  = L2

    # Self-weight: mass per mm of member length [kg/mm]
    # rho [kg/m³] × A [mm²] × (1e-3 m/mm)² / (1e-3 m/mm) = rho × A × 1e-9 [kg/mm]
    sw_kg_per_mm = rho * A * 1e-9   # kg/mm

    # Build the model
    try:
        frame = FEModel3D()

        frame.add_node('N1',  0.0,    0.0,   0.0)
        frame.add_node('Nm',  x_mid,  0.0,   0.0)   # midpoint of M1, sensor location
        frame.add_node('N2',  x_n2,   0.0,   0.0)
        frame.add_node('N3',  x_n3,   y_n3,  0.0)

        frame.def_support('N1', True, True, True, True, True, True)

        # Material: nu=0.3 (typical metal), rho not used (self-weight applied manually)
        frame.add_material('mat', E, G, 0.3, 0.0)

        # Section: (name, A, Iy, Iz, J) — all in mm², mm⁴
        frame.add_section('sec', A, Iy, Iz, J)

        frame.add_member('M1a', 'N1', 'Nm', 'mat', 'sec')
        frame.add_member('M1b', 'Nm', 'N2', 'mat', 'sec')
        frame.add_member('M2',  'N2', 'N3', 'mat', 'sec')

        member_names = ['M1a', 'M1b', 'M2']
        tc_mass = TORCH_MASS + CABLE_MASS   # kg — both act at N3

        # Apply loads for each gravity case
        for c in cases:
            cn = c['name']
            gx, gy, gz = c['gx'], c['gy'], c['gz']

            # Point loads: torch + cable at N3 [N = kg × m/s²]
            if abs(gx) > 1e-12: frame.add_node_load('N3', 'FX', tc_mass * gx, case=cn)
            if abs(gy) > 1e-12: frame.add_node_load('N3', 'FY', tc_mass * gy, case=cn)
            if abs(gz) > 1e-12: frame.add_node_load('N3', 'FZ', tc_mass * gz, case=cn)

            # Point loads: sensor at Nm [N]
            if abs(gx) > 1e-12: frame.add_node_load('Nm', 'FX', SENSOR_MASS * gx, case=cn)
            if abs(gy) > 1e-12: frame.add_node_load('Nm', 'FY', SENSOR_MASS * gy, case=cn)
            if abs(gz) > 1e-12: frame.add_node_load('Nm', 'FZ', SENSOR_MASS * gz, case=cn)

            # Distributed self-weight [N/mm] — global directions FX/FY/FZ
            # sw_kg_per_mm [kg/mm] × g_component [m/s²] → [N/mm]
            fx_sw = sw_kg_per_mm * gx
            fy_sw = sw_kg_per_mm * gy
            fz_sw = sw_kg_per_mm * gz
            for mn in member_names:
                if abs(fx_sw) > 1e-12: frame.add_member_dist_load(mn, 'FX', fx_sw, fx_sw, case=cn)
                if abs(fy_sw) > 1e-12: frame.add_member_dist_load(mn, 'FY', fy_sw, fy_sw, case=cn)
                if abs(fz_sw) > 1e-12: frame.add_member_dist_load(mn, 'FZ', fz_sw, fz_sw, case=cn)

        # One combo per case (factor = 1.0 — loads already include the load factor)
        for c in cases:
            frame.add_load_combo(c['name'], {c['name']: 1.0})

        frame.analyze(check_stability=False, log=False)

    except Exception as e:
        return None, str(e)

    # ── Extract results ───────────────────────────────────────────────────────
    try:
        # Section moduli [mm³]
        S_strong = Iy / (D / 2)
        S_weak   = Iz / (W / 2)
        S_min    = min(S_strong, S_weak)

        # Member masses [kg]
        m_m1a = member_mass_kg(rho, A, L_m1a)
        m_m1b = member_mass_kg(rho, A, L_m1b)
        m_m2  = member_mass_kg(rho, A, L_m2)
        frame_mass = m_m1a + m_m1b + m_m2
        total_mass = frame_mass + TORCH_MASS + SENSOR_MASS + CABLE_MASS + FASTENER_MASS

        # Worst-case trackers (across dynamic load cases)
        worst_tcp_defl      = 0.0
        worst_sensor_defl   = 0.0
        worst_bending_stress = 0.0
        worst_Mz            = 0.0
        worst_My            = 0.0
        worst_shear         = 0.0
        worst_axial         = 0.0
        worst_case_name     = ''
        worst_alpha         = None

        for c in cases:
            if not c['dynamic']:
                continue
            cn = c['name']

            # Node displacements [mm]
            dx3 = frame.nodes['N3'].DX.get(cn, 0.0)
            dy3 = frame.nodes['N3'].DY.get(cn, 0.0)
            dz3 = frame.nodes['N3'].DZ.get(cn, 0.0)
            tcp_defl = math.sqrt(dx3**2 + dy3**2 + dz3**2)

            dxm = frame.nodes['Nm'].DX.get(cn, 0.0)
            dym = frame.nodes['Nm'].DY.get(cn, 0.0)
            dzm = frame.nodes['Nm'].DZ.get(cn, 0.0)
            sensor_defl = math.sqrt(dxm**2 + dym**2 + dzm**2)

            # Member forces — use max/min envelope methods
            case_Mz = case_My = case_shear = case_axial = 0.0
            for mn in member_names:
                mem = frame.members[mn]
                case_Mz    = max(case_Mz,    abs(mem.max_moment('Mz', cn)),
                                              abs(mem.min_moment('Mz', cn)))
                case_My    = max(case_My,    abs(mem.max_moment('My', cn)),
                                              abs(mem.min_moment('My', cn)))
                case_shear = max(case_shear, abs(mem.max_shear('Fy', cn)),
                                              abs(mem.min_shear('Fy', cn)),
                                              abs(mem.max_shear('Fz', cn)),
                                              abs(mem.min_shear('Fz', cn)))
                case_axial = max(case_axial, abs(mem.max_axial(cn)),
                                              abs(mem.min_axial(cn)))

            # Bending stress at extreme fiber [MPa = N/mm²]
            M_max = max(case_Mz, case_My)
            bending_stress = M_max / S_min if S_min > 0 else 0.0

            if tcp_defl > worst_tcp_defl:
                worst_tcp_defl  = tcp_defl
                worst_case_name = cn
                worst_alpha     = c['alpha']
            worst_sensor_defl    = max(worst_sensor_defl, sensor_defl)
            worst_bending_stress = max(worst_bending_stress, bending_stress)
            worst_Mz             = max(worst_Mz, case_Mz)
            worst_My             = max(worst_My, case_My)
            worst_shear          = max(worst_shear, case_shear)
            worst_axial          = max(worst_axial, case_axial)

        stress_util = worst_bending_stress / Fy

        # ── Wrist load calculations ───────────────────────────────────────────
        # Weighted CG from Node 1 [mm]
        cg_m1a_x, cg_m1a_y = L_m1a / 2, 0.0
        cg_m1b_x, cg_m1b_y = x_mid + L_m1b / 2, 0.0
        cg_m2_x  = (x_n2 + x_n3) / 2
        cg_m2_y  = y_n3 / 2
        masses_cg = [
            (m_m1a,                  cg_m1a_x, cg_m1a_y),
            (m_m1b,                  cg_m1b_x, cg_m1b_y),
            (m_m2,                   cg_m2_x,  cg_m2_y),
            (TORCH_MASS + CABLE_MASS, x_n3,     y_n3),
            (SENSOR_MASS,             x_mid,    0.0),
            (FASTENER_MASS,           x_n3/2,   y_n3/4),
        ]
        cg_x_mm = sum(m * x for m, x, y in masses_cg) / total_mass
        cg_y_mm = sum(m * y for m, x, y in masses_cg) / total_mass

        cg_r_m = math.sqrt(cg_x_mm**2 + cg_y_mm**2) / 1000.0  # radial [m]
        cg_x_m = cg_x_mm / 1000.0                               # X offset [m]

        j45_moment_Nm   = total_mass * G_ACCEL * cg_r_m
        j6_moment_Nm    = total_mass * G_ACCEL * cg_x_m * 0.3   # eccentricity estimate
        j45_inertia_kgm2 = total_mass * cg_r_m**2
        j6_inertia_kgm2  = total_mass * cg_x_m**2

        # ── Pass/Fail checks ──────────────────────────────────────────────────
        checks = {
            'payload':     total_mass        <= ROBOT_LIMITS['payload_kg'],
            'j45_moment':  j45_moment_Nm     <= ROBOT_LIMITS['j45_moment_Nm'],
            'j6_moment':   j6_moment_Nm      <= ROBOT_LIMITS['j6_moment_Nm'],
            'j45_inertia': j45_inertia_kgm2  <= ROBOT_LIMITS['j45_inertia_kgm2'],
            'j6_inertia':  j6_inertia_kgm2   <= ROBOT_LIMITS['j6_inertia_kgm2'],
            'deflection':  worst_tcp_defl    <= MAX_DEFLECTION,
            'stress':      stress_util       <= 1.0,
        }

        return {
            'material':            mat_name,
            'D_mm':                D,
            'W_mm':                W,
            't_mm':                t,
            'A_mm2':               A,
            'Iy_mm4':              Iy,
            'Iz_mm4':              Iz,
            'J_mm4':               J,
            'frame_mass_kg':       frame_mass,
            'total_mass_kg':       total_mass,
            'tcp_defl_mm':         worst_tcp_defl,
            'sensor_defl_mm':      worst_sensor_defl,
            'bending_stress_MPa':  worst_bending_stress,
            'stress_util_pct':     stress_util * 100,
            'Mz_max_Nmm':          worst_Mz,
            'My_max_Nmm':          worst_My,
            'shear_max_N':         worst_shear,
            'axial_max_N':         worst_axial,
            'worst_case':          worst_case_name,
            'worst_alpha':         worst_alpha,
            'cg_x_mm':             cg_x_mm,
            'cg_y_mm':             cg_y_mm,
            'j45_moment_Nm':       j45_moment_Nm,
            'j6_moment_Nm':        j6_moment_Nm,
            'j45_inertia_kgm2':    j45_inertia_kgm2,
            'j6_inertia_kgm2':     j6_inertia_kgm2,
            'j45_util_pct':        j45_moment_Nm  / ROBOT_LIMITS['j45_moment_Nm']  * 100,
            'j6_util_pct':         j6_moment_Nm   / ROBOT_LIMITS['j6_moment_Nm']   * 100,
            'checks':              checks,
            'passes':              all(checks.values()),
        }, None

    except Exception as e:
        return None, f'Result extraction: {e}'


# ─────────────────────────────────────────────────────────────────────────────
# Console output
# ─────────────────────────────────────────────────────────────────────────────

COL_W = 120

def print_table(results, title, max_rows=20, sort_key='total_mass_kg', reverse=False):
    if not results:
        print(f'\n{title}: (no results)')
        return
    rows = sorted(results, key=lambda r: r[sort_key], reverse=reverse)[:max_rows]
    print(f'\n{"═" * COL_W}')
    print(f'  {title}')
    print(f'{"═" * COL_W}')
    print(
        f'{"Material":<14} {"D":>6} {"W":>6} {"t":>6}  '
        f'{"Frame kg":>9} {"Total kg":>9}  '
        f'{"TCP Defl":>9} {"Stress%":>8}  '
        f'{"Worst Ang":>9}  '
        f'{"J4/5 N·m":>9} {"J4/5 Util%":>11}  '
        f'{"J6 N·m":>8}  {"PASS/FAIL":>9}'
    )
    print('─' * COL_W)
    for r in rows:
        wa = f"{r['worst_alpha']}°" if r['worst_alpha'] is not None else 'Z-axis'
        flag = 'PASS' if r['passes'] else 'FAIL'
        # Show which check(s) failed
        if not r['passes']:
            fails = [k for k, v in r['checks'].items() if not v]
            flag += f" ({','.join(fails)})"
        print(
            f"{r['material']:<14} {r['D_mm']:>6.1f} {r['W_mm']:>6.1f} {r['t_mm']:>6.2f}  "
            f"{r['frame_mass_kg']:>9.3f} {r['total_mass_kg']:>9.3f}  "
            f"{r['tcp_defl_mm']:>9.4f} {r['stress_util_pct']:>7.1f}%  "
            f"{wa:>9}  "
            f"{r['j45_moment_Nm']:>9.2f} {r['j45_util_pct']:>10.1f}%  "
            f"{r['j6_moment_Nm']:>8.2f}  {flag}"
        )


# ─────────────────────────────────────────────────────────────────────────────
# Excel output
# ─────────────────────────────────────────────────────────────────────────────

def save_excel(all_results, filename='eoat_sweep_results.xlsx'):
    if not EXCEL_AVAILABLE:
        print('Skipping Excel output (openpyxl not available).')
        return

    wb = openpyxl.Workbook()

    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill   = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    hdr_fill   = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    hdr_font   = Font(color='FFFFFF', bold=True)

    COLS = [
        'Material', 'D (mm)', 'W (mm)', 't (mm)',
        'A (mm²)', 'Iy (mm⁴)', 'Iz (mm⁴)', 'J (mm⁴)',
        'Frame Mass (kg)', 'Total Mass (kg)',
        'Max TCP Defl (mm)', 'Max Sensor Defl (mm)',
        'Max Bending Stress (MPa)', 'Stress Util (%)',
        'Mz Max (N·mm)', 'My Max (N·mm)',
        'Shear Max (N)', 'Axial Max (N)',
        'Worst Gravity Case', 'Worst Angle (°)',
        'CG X (mm)', 'CG Y (mm)',
        'J4/5 Moment (N·m)', 'J6 Moment (N·m)',
        'J4/5 Inertia (kg·m²)', 'J6 Inertia (kg·m²)',
        'J4/5 Util (%)', 'J6 Util (%)',
        'Payload OK', 'J4/5 Moment OK', 'J6 Moment OK',
        'J4/5 Inertia OK', 'J6 Inertia OK',
        'Deflection OK', 'Stress OK',
        'PASS/FAIL',
    ]

    def write_sheet(ws, results):
        ws.append(COLS)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal='center')

        for r in results:
            wa = str(r['worst_alpha']) if r['worst_alpha'] is not None else 'Z'
            ch = r['checks']
            ws.append([
                r['material'],
                round(r['D_mm'], 1),      round(r['W_mm'], 1),      round(r['t_mm'], 2),
                round(r['A_mm2'], 2),     round(r['Iy_mm4'], 0),    round(r['Iz_mm4'], 0),
                round(r['J_mm4'], 0),
                round(r['frame_mass_kg'],  4),  round(r['total_mass_kg'], 4),
                round(r['tcp_defl_mm'],    5),  round(r['sensor_defl_mm'], 5),
                round(r['bending_stress_MPa'], 2), round(r['stress_util_pct'], 2),
                round(r['Mz_max_Nmm'],  1),    round(r['My_max_Nmm'],  1),
                round(r['shear_max_N'], 2),    round(r['axial_max_N'], 2),
                r['worst_case'], wa,
                round(r['cg_x_mm'], 2),        round(r['cg_y_mm'], 2),
                round(r['j45_moment_Nm'],    3), round(r['j6_moment_Nm'],   3),
                round(r['j45_inertia_kgm2'], 4), round(r['j6_inertia_kgm2'], 4),
                round(r['j45_util_pct'], 1),   round(r['j6_util_pct'], 1),
                'YES' if ch['payload']     else 'NO',
                'YES' if ch['j45_moment']  else 'NO',
                'YES' if ch['j6_moment']   else 'NO',
                'YES' if ch['j45_inertia'] else 'NO',
                'YES' if ch['j6_inertia']  else 'NO',
                'YES' if ch['deflection']  else 'NO',
                'YES' if ch['stress']      else 'NO',
                'PASS' if r['passes'] else 'FAIL',
            ])
            fill = green_fill if r['passes'] else red_fill
            for cell in ws[ws.max_row]:
                cell.fill = fill

        for i in range(1, len(COLS) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 18

    ws1 = wb.active
    ws1.title = 'All Results'
    write_sheet(ws1, sorted(all_results, key=lambda r: r['total_mass_kg']))

    ws2 = wb.create_sheet('Passing Designs')
    passing = [r for r in all_results if r['passes']]
    write_sheet(ws2, sorted(passing, key=lambda r: r['total_mass_kg']))

    wb.save(filename)
    print(f'Excel saved: {filename}')


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    t_start = time.time()

    # Build combination list
    combos = [
        (mat, D, W, t)
        for mat in MATERIALS
        for D in D_VALUES
        for W in W_VALUES
        if W <= D
        for t in T_VALUES
        if t < D / 4 and t < W / 4
    ]

    cases = gravity_cases()
    total = len(combos)

    print(f'\nEOAT Parametric Sweep')
    print('=' * 60)
    print(f'Geometry  : L1={L1} mm, L2={L2} mm, θ={THETA}°')
    print(f'Masses    : Torch={TORCH_MASS} kg, Sensor={SENSOR_MASS} kg, '
          f'Cable={CABLE_MASS} kg, Fasteners={FASTENER_MASS} kg')
    print(f'Dyn factor: {G_FACTOR}×    Defl limit: {MAX_DEFLECTION} mm')
    print(f'Combos    : {total} sections × {len(cases)} load cases = {total * len(cases)} analyses')
    print('=' * 60)

    all_results = []
    failures = []

    for i, (mat_name, D, W, t) in enumerate(combos, 1):
        label = f'{mat_name}  {D:.1f}×{W:.1f}×{t:.2f} mm'
        print(f'\rAnalyzing {i:>4}/{total}: {label:<45}', end='', flush=True)
        result, err = run_analysis(mat_name, MATERIALS[mat_name], D, W, t, cases)
        if result is None:
            failures.append((label, err))
        else:
            all_results.append(result)

    print()  # end progress line

    passing = [r for r in all_results if r['passes']]
    failing  = [r for r in all_results if not r['passes']]

    print(f'\nResults: {len(all_results)} analyzed  |  '
          f'{len(passing)} PASS  |  {len(failing)} FAIL  |  {len(failures)} errored\n')

    print_table(passing,
                f'TOP 20 LIGHTEST PASSING DESIGNS  (sorted by total mass)',
                max_rows=20, sort_key='total_mass_kg')

    print_table(passing,
                f'TOP 5 STIFFEST PASSING DESIGNS  (sorted by TCP deflection)',
                max_rows=5, sort_key='tcp_defl_mm')

    if failures:
        print(f'\n{"─" * 60}')
        print(f'Analysis failures ({len(failures)}):')
        for label, err in failures[:10]:
            print(f'  {label}: {err}')
        if len(failures) > 10:
            print(f'  … and {len(failures) - 10} more')

    save_excel(all_results)

    elapsed = time.time() - t_start
    print(f'\nRuntime: {elapsed:.1f} s')
    print('Done.\n')


if __name__ == '__main__':
    main()
